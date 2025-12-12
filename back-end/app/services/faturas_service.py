from typing import List, Tuple
from sqlalchemy.orm import Session
from app.infrastructure.orm_models import TarifaAnoDB, TarifaFaixaDB, ImovelDB
from app.models.schemas import TarifaAnoCreate, TarifaFaixa, FaturaCalculoRequest, FaturaCalculoResponse, CategoriaLigacao


def upsert_tarifa_ano(db: Session, tarifa: TarifaAnoCreate) -> int:
    existing = db.query(TarifaAnoDB).filter(
        TarifaAnoDB.ano == tarifa.ano,
        TarifaAnoDB.tipo == tarifa.tipo_imovel,
        TarifaAnoDB.categoria_ligacao == tarifa.categoria_ligacao,
    ).first()
    if existing is None:
        existing = TarifaAnoDB(ano=tarifa.ano, tipo=tarifa.tipo_imovel, categoria_ligacao=tarifa.categoria_ligacao)
        db.add(existing)
        db.flush()
    # remove faixas anteriores
    db.query(TarifaFaixaDB).filter(TarifaFaixaDB.tarifa_ano_id == existing.id).delete()
    # inserir faixas novas
    for f in tarifa.faixas:
        db.add(TarifaFaixaDB(
            tarifa_ano_id=existing.id,
            faixa=f.faixa,
            consumo_min=f.consumo_min,
            consumo_max=f.consumo_max,
            valor_minimo=int(f.valor_minimo * 100) if f.valor_minimo is not None else None,
            valor_por_m3=int(f.valor_por_m3 * 100) if f.valor_por_m3 is not None else None,
        ))
    db.commit()
    return existing.id


def calcular_fatura(db: Session, req: FaturaCalculoRequest) -> FaturaCalculoResponse:
    imovel = db.query(ImovelDB).filter(ImovelDB.matricula == req.matricula_imovel).first()
    if not imovel:
        raise ValueError("Imóvel não encontrado")
    categoria = imovel.categoria
    tarifa_ano = db.query(TarifaAnoDB).filter(
        TarifaAnoDB.ano == req.ano,
        TarifaAnoDB.tipo == imovel.tipo,
        TarifaAnoDB.categoria_ligacao == categoria,
    ).first()
    if not tarifa_ano:
        raise ValueError("Tarifas não cadastradas para o ano e tipo do imóvel")
    faixas = sorted(tarifa_ano.faixas, key=lambda x: x.faixa)
    consumo_restante = req.consumo_m3
    valor_agua_cent = 0
    detalhes: List[dict] = []
    for f in faixas:
        faixa_consumo = 0
        if f.faixa == 1:
            # tarifa mínima até 6 m3
            faixa_consumo = min(consumo_restante, f.consumo_max or consumo_restante)
            if f.valor_minimo is not None:
                valor_agua_cent += f.valor_minimo
                detalhes.append({"faixa": f.faixa, "m3": faixa_consumo, "valor": f.valor_minimo / 100.0, "tipo": "mínimo"})
            if consumo_restante <= (f.consumo_max or 0):
                consumo_restante = 0
            else:
                consumo_restante -= (f.consumo_max or 0)
        else:
            limite_superior = f.consumo_max or req.consumo_m3
            capacidade = max(0, limite_superior - f.consumo_min + 1)
            faixa_consumo = min(consumo_restante, capacidade)
            if faixa_consumo > 0 and f.valor_por_m3 is not None:
                valor = faixa_consumo * f.valor_por_m3
                valor_agua_cent += valor
                detalhes.append({"faixa": f.faixa, "m3": faixa_consumo, "valor": valor / 100.0, "tipo": "por_m3"})
            consumo_restante -= faixa_consumo
        if consumo_restante <= 0:
            break
    # esgoto: regra simplificada 80% da água se imovel.esgoto True
    valor_esgoto_cent = int(valor_agua_cent * 0.8) if imovel.esgoto else 0
    total_cent = valor_agua_cent + valor_esgoto_cent
    return FaturaCalculoResponse(
        ano=req.ano,
        matricula_imovel=req.matricula_imovel,
        consumo_m3=req.consumo_m3,
        valor_agua=valor_agua_cent / 100.0,
        valor_esgoto=valor_esgoto_cent / 100.0,
        total=total_cent / 100.0,
        detalhamento=detalhes,
    )


def carregar_tarifas_xlsx(db: Session, caminho_xlsx: str, ano: int, tipo_imovel, categoria_ligacao: CategoriaLigacao) -> int:
    from openpyxl import load_workbook
    wb = load_workbook(caminho_xlsx)
    ws = wb.active
    faixas: List[TarifaFaixa] = []
    # Expectativa: colunas: faixa, consumo_min, consumo_max, valor_minimo, valor_por_m3
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        faixa, cmin, cmax, vmin, vpm3 = row
        faixas.append(TarifaFaixa(
            faixa=int(faixa),
            consumo_min=int(cmin),
            consumo_max=int(cmax) if cmax is not None else None,
            valor_minimo=float(vmin) if vmin is not None else None,
            valor_por_m3=float(vpm3) if vpm3 is not None else None,
        ))
    tid = upsert_tarifa_ano(db, TarifaAnoCreate(ano=ano, tipo_imovel=tipo_imovel, categoria_ligacao=categoria_ligacao, faixas=faixas))
    return tid