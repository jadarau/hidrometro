import json
import sys
from typing import Dict, List, Any

try:
    import openpyxl  # type: ignore
except Exception as e:
    print("[erro] openpyxl não instalado no venv. Instale com 'pip install openpyxl'", file=sys.stderr)
    raise


HEADER_MAP = {
    # possíveis nomes de colunas (normalizados para minúsculas e sem acentos)
    "tipo_imovel": ["tipo_imovel", "tipo de imovel", "tipo"],
    "categoria_ligacao": ["categoria_ligacao", "categoria de ligacao", "categoria", "ligacao"],
    "faixa": ["faixa", "grupo", "tier"],
    "consumo_min": ["consumo_min", "min", "consumo minimo"],
    "consumo_max": ["consumo_max", "max", "consumo maximo"],
    "valor_minimo": ["valor_minimo", "valor minimo", "tarifa minima"],
    "valor_por_m3": ["valor_por_m3", "valor por m3", "preco m3", "preco por m3"],
}


def normalize(s: str) -> str:
    import unicodedata
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.strip().lower()


def match_headers(row_values: List[Any]) -> Dict[str, int]:
    idx: Dict[str, int] = {}
    for i, val in enumerate(row_values):
        if val is None:
            continue
        key = normalize(str(val))
        for target, aliases in HEADER_MAP.items():
            if key in aliases and target not in idx:
                idx[target] = i
    return idx


def parse_workbook(path: str) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    tarifas_rows: List[Dict[str, Any]] = []
    for ws in wb.worksheets:
        # tenta localizar cabeçalho na primeira ou segunda linha
        header_row = None
        for header_try in (1, 2, 3):
            header_values = [ws.cell(row=header_try, column=c).value for c in range(1, ws.max_column + 1)]
            headers = match_headers(header_values)
            if "tipo_imovel" in headers and "categoria_ligacao" in headers:
                header_row = header_try
                header_idx = headers
                break
        if not header_row:
            continue

        for r in range(header_row + 1, ws.max_row + 1):
            def get(colname: str):
                ci = header_idx.get(colname)
                return ws.cell(row=r, column=(ci + 1)).value if ci is not None else None

            tipo_imovel = get("tipo_imovel")
            categoria_ligacao = get("categoria_ligacao")
            if not tipo_imovel or not categoria_ligacao:
                continue
            faixa = get("faixa")
            consumo_min = get("consumo_min")
            consumo_max = get("consumo_max")
            valor_minimo = get("valor_minimo")
            valor_por_m3 = get("valor_por_m3")

            try:
                consumo_min = int(consumo_min) if consumo_min is not None and str(consumo_min).strip() != "" else 0
            except Exception:
                consumo_min = 0
            try:
                consumo_max = int(consumo_max) if consumo_max is not None and str(consumo_max).strip() != "" else None
            except Exception:
                consumo_max = None
            try:
                valor_minimo = float(valor_minimo) if valor_minimo is not None and str(valor_minimo).strip() != "" else None
            except Exception:
                valor_minimo = None
            try:
                valor_por_m3 = float(valor_por_m3) if valor_por_m3 is not None and str(valor_por_m3).strip() != "" else None
            except Exception:
                valor_por_m3 = None

            row = {
                "tipo_imovel": str(tipo_imovel).strip(),
                "categoria_ligacao": str(categoria_ligacao).strip(),
                "faixa": int(faixa) if isinstance(faixa, (int, float)) else 1,
                "consumo_min": consumo_min,
                "consumo_max": consumo_max,
                "valor_minimo": valor_minimo,
                "valor_por_m3": valor_por_m3,
            }
            tarifas_rows.append(row)
    return tarifas_rows


def group_tarifas(rows: List[Dict[str, Any]], ano: int) -> Dict[str, Any]:
    # agrupa por (tipo_imovel, categoria_ligacao)
    agrupado: Dict[tuple, List[Dict[str, Any]]] = {}
    for r in rows:
        key = (r["tipo_imovel"], r["categoria_ligacao"])
        agrupado.setdefault(key, []).append(r)

    tarifas: List[Dict[str, Any]] = []
    for (tipo_imovel, categoria_ligacao), items in agrupado.items():
        faixas = []
        for it in items:
            faixas.append({
                "faixa": it["faixa"],
                "consumo_min": it["consumo_min"],
                "consumo_max": it["consumo_max"],
                "valor_minimo": it["valor_minimo"],
                "valor_por_m3": it["valor_por_m3"],
            })
        tarifas.append({
            "ano": ano,
            "tipo_imovel": tipo_imovel,
            "categoria_ligacao": categoria_ligacao,
            "faixas": faixas,
        })
    return {"tarifas": tarifas}


def main():
    if len(sys.argv) < 3:
        print("Uso: python -m app.tools.export_tarifas_parse_xlsx <caminho_xlsx> <ano>", file=sys.stderr)
        sys.exit(1)
    caminho = sys.argv[1]
    ano = int(sys.argv[2])
    rows = parse_workbook(caminho)
    if not rows:
        print("[erro] Nenhuma linha de tarifas encontrada no XLSX.", file=sys.stderr)
    payload = group_tarifas(rows, ano)
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()